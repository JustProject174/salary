import requests
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import io
import matplotlib

matplotlib.use('Agg')  # Устанавливаем неинтерактивный бэкенд


def parse_vacancies_api(job_title, pages=5):
    vacancies = []
    url = "https://api.hh.ru/vacancies"

    for page in range(pages):
        params = {
            "text": job_title,
            "area": 113,
            "page": page,
            "per_page": 20
        }

        response = requests.get(url, params=params)

        if response.status_code != 200:
            print(f"Ошибка при запросе на страницу {page + 1}: {response.status_code}")
            continue

        data = response.json()

        for vacancy in data['items']:
            title = vacancy.get('name', 'Нет названия')
            salary = vacancy.get('salary', None)
            if salary:
                salary = parse_salary(salary)

            if vacancy['area']['id'] != "1":
                vacancies.append({
                    "title": title,
                    "salary": salary,
                    "city": vacancy['area']['name'],
                    "company": vacancy.get('employer', {}).get('name', 'Не указана')
                })

    print(f"Найдено вакансий: {len(vacancies)}")
    return vacancies


def parse_salary(salary_data):
    if salary_data['from'] and salary_data['to']:
        return (salary_data['from'] + salary_data['to']) // 2
    elif salary_data['from']:
        return salary_data['from']
    elif salary_data['to']:
        return salary_data['to']
    return None


def save_to_excel_with_summary(vacancies, filename="vacancies_data.xlsx"):
    # Создаем DataFrame
    df = pd.DataFrame(vacancies)

    # Рассчитываем статистику
    median_salary = df['salary'].median()
    p25 = df['salary'].quantile(0.25)
    p75 = df['salary'].quantile(0.75)

    # Создаем сводную таблицу
    pivot = df.groupby('city')['salary'].agg(['median', 'mean', 'count'])
    pivot = pivot.rename(columns={
        'median': 'Медиана',
        'mean': 'Средняя',
        'count': 'Количество'
    }).reset_index()

    # Создаем график
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.scatter(df['salary'], df['city'], alpha=0.5)
    ax.axvline(median_salary, color='r', linestyle='--', label=f'Медиана: {median_salary:.0f} руб.')
    ax.set_title("Распределение зарплат по городам")
    ax.set_xlabel("Зарплата (руб.)")
    ax.set_ylabel("Город")
    ax.legend()
    plt.tight_layout()

    # Сохраняем график в буфер
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=300)
    img_buffer.seek(0)
    plt.close()

    # Создаем Excel-файл
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Вакансии"

    # Записываем данные вакансий
    for r_idx, row in enumerate(pd.DataFrame(vacancies).itertuples(), 1):
        for c_idx, value in enumerate(row[1:], 1):
            ws1.cell(row=r_idx, column=c_idx, value=value)

    # Добавляем сводную таблицу
    ws2 = wb.create_sheet("Статистика")
    for r_idx, row in enumerate(pivot.itertuples(), 1):
        for c_idx, value in enumerate(row[1:], 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    # Добавляем график
    ws3 = wb.create_sheet("График")
    img = Image(img_buffer)
    ws3.add_image(img, 'A1')

    # Настраиваем ширину столбцов
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    # Сохраняем файл
    wb.save(filename)
    print(f"Данные сохранены в файл: {filename}")


def main():
    job_title = input("Введите должность: ")
    vacancies = parse_vacancies_api(job_title)

    if vacancies:
        save_to_excel_with_summary(vacancies)
        print("Анализ завершен успешно!")
    else:
        print(f"Не удалось найти вакансии для должности '{job_title}'")


if __name__ == "__main__":
    main()